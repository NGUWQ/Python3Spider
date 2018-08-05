#提取姓名
import openpyxl
import json
wbr=openpyxl.load_workbook('C:\\Users\\TTT\\Desktop\\杂\\1.xlsx')
ws=wbr['15计科本1']
lst=[]
for cell in list(ws.rows)[20]:
    lst.append(cell.value)
    lst0=[]
for cell in list(ws.rows)[0]:
    lst0.append(cell.value)

wbw=openpyxl.Workbook()
wbw.remove(wbw.active)
nws=wbw.create_sheet('mysheet',index=1)
nws.append(lst)
for i in range(10):
    nws['C%d'%(i+2)].value=i+1
    nws['b%d'%(i+2)].value=i+1
    nws['D11'].value='=sum(A2:A11)'
    wbw.save('C:\\Users\\TTT\\Desktop\\杂\\2.xlsx')